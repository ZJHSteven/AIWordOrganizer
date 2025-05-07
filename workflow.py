from LlmCompletion import Agent # 从 LlmCompletion.py 导入 LlmApiAgent 类
from LlmCompletion import state

# 创建 LlmApiAgent 实例，传入模型名称
# 您可以根据需要更改模型名称

agent1_extractor = Agent(
    name="english_chinese_pair_extractor_agent",
    model="gemini/gemini-2.0-flash",
    description="从系统解剖学文本中提取英文术语和对应的中文含义，并以 Markdown 有序列表格式输出。",
    instruction=(
        "你的任务是从用户提供的【系统解剖学】相关文本中，提取出成对的中文术语及其对应的英文单词。 "
        "文本的格式可能多样，例如 '中文术语 English Word'，或者混合在段落中。 "
        "请仔细阅读文本，找出所有这样的配对。 "
        "最终，你必须只输出一个 Markdown 格式的有序列表，其中每一项都是中文术语和括号括起来的英文单词。 "
        "绝对不要包含任何额外的解释性文字、Markdown 代码块标记（例如 \\`\\`\\`）或其他任何前缀或后缀。 "
        "输出格式范例如下：\\n"
        "1. 肱骨 (Humerus)\\n"
        "2. 踝 (Ankle)\\n"
        "如果找不到任何配对，请输出一个空字符串或提示信息，例如“未找到可提取的术语。”。"
    ),
    tools=[],
    output_key="extracted_pairs"
)


# --- Agent 2 Definition (Decomposer) ---
agent2_decomposer = Agent(
    name="word_decomposer_agent",
    model="gemini/gemini-2.0-flash",
    description="接收英汉词对字典，对英文单词进行词根拆解，并以 Markdown 有序列表格式化输出。",
    instruction=(
        "你将接收一个系统解剖学单词表 {extracted_pairs} 的 Python 字典作为输入，其中键是英文单词，值是对应的中文释义。\\n"
        "你的任务是：对每一个英文单词进行详细的词根、词缀分析，尽可能追溯到拉丁语或希腊语等原始构词来源，并结合中文释义进行说明。不要简单以词组形式解释整个单词，而是进一步拆解为词根和词缀组成。如果原始构词与解剖学术语意思差别较大，可在中文释义部分加以说明。\\n"
        "最终，你必须只输出一个 Markdown 格式的有序列表。列表中的每一项代表一个单词的分析结果。\\n"
        "每一项的格式如下，请严格遵守（注意换行和缩进）：\\n"
        "1. EnglishWord\\n"
        "   释义：中文意思\\n"
        "   词根：“root1-” (来源和含义)\\n"
        "   词根：“root2-” (来源和含义)\\n"
        "   后缀：“-suffix” (来源和含义)\\n"
        "\\n"
        "输出格式范例如下，假设输入为 `{\\'pilosebaceous\\': \\'毛囊皮脂腺的\\', \\'cardiopathy\\': \\'心脏病\\'}`：\\n"
        "1. pilosebaceous\\n"
        "   释义：毛囊皮脂腺的\\n"
        "   词根：“pilo-” (毛发，来自拉丁语“pilus”)\\n"
        "   词根：“sebace-” (皮脂的，来自拉丁语“sebaceus”)\\n"
        "   后缀：“-ous” (形容词后缀，表示“具有……的”)\\n"
        "2. cardiopathy\\n"
        "   释义：心脏病\\n"
        "   词根：“cardio-” (心脏，来自希腊语“kardia”)\\n"
        "   词根：“-pathy” (疾病，来自希腊语“pathos”)\\n"
        "\\n"
        "请确保输出的是纯净的 Markdown 有序列表。\\n"
        "重要：输出结果的开头和结尾绝对不能包含三个反引号（\\`\\`\\`）或其他任何形式的代码块标记。输出必须直接以有序列表的第一项开始（例如 '1. EnglishWord...'）。不要添加任何前缀、后缀或解释性文字。\\n"
        "如果输入的字典为空，请输出一个空字符串或提示信息，例如“没有可供分析的单词。”。"
    ),
    tools=[],
    output_key="decomposed_markdown_list"
)

def process_markdown_and_insert_to_excel():
    import re
    import openpyxl

    # 1. 提取 Markdown 字符串
    markdown_text = state.get('decomposed_markdown_list', '')

    # 如果存在 Markdown 代码块标记，则移除
    if markdown_text.startswith("```markdown\n"):
        markdown_text = markdown_text[len("```markdown\n"):]
    if markdown_text.endswith("\n```"):
        markdown_text = markdown_text[:-len("\n```")]
    markdown_text = markdown_text.strip()  # 移除首尾可能存在的空白

    # 2. 定义正则表达式
# 匹配 "数字. 单词\n" 和随后的 "   释义：... (直到下一个 "数字." 或字符串末尾)"
# (?s) 使 . 可以匹配换行符
# (\d+\.\s+.*?)\n  匹配 "数字. 单词" 这一行，捕获单词
# (\s{3}释义：.*?)(?=\n\d+\.\s+|\Z) 匹配 "   释义：" 开始的详细内容，直到下一个条目或字符串末尾
    regex = r"(\d+\.\s*(.*?)\n)(\s+释义：.*?)(?=\n\d+\.\s*|\Z)"

    parsed_data = []
    if markdown_text:
        matches = re.finditer(regex, markdown_text, re.DOTALL)
        for match in matches:
            english_word = match.group(2).strip()

            chinese_definition_and_roots = match.group(3).strip()
            chinese_definition_and_roots = chinese_definition_and_roots.replace('   ', '')  # 去除多余的空格
            chinese_definition_and_roots = chinese_definition_and_roots.replace('    ', '')

            parsed_data.append({
                "英文": english_word,
                "中文": chinese_definition_and_roots
            })

    # 3. 准备 Excel 文件路径和表头
    excel_file_name = "test_output.xlsx"  # 测试文件名
# target_excel_path = "d:\Workspace\Stable\Python\ReWordOrganizer\词根整理.xlsx" # 实际文件路径（注释掉供测试）
    excel_column_english = "单词"
    excel_column_chinese = "含义（可不填）"
    excel_column_tag = "标签（可不填，多个标签请用英文逗号分隔）"

    # 打开 Excel 文件
    wb = openpyxl.load_workbook(excel_file_name)
    sheet = wb.active

    # 获取表头
    headers = [cell.value for cell in sheet[1]]

    # 查找列索引
    english_col_idx = headers.index(excel_column_english) + 1 if excel_column_english in headers else None
    chinese_col_idx = headers.index(excel_column_chinese) + 1 if excel_column_chinese in headers else None
    tag_col_idx = headers.index(excel_column_tag) + 1 if excel_column_tag in headers else None

    # 如果表头中没有找到列，则添加表头
    if english_col_idx is None:
        english_col_idx = len(headers) + 1
        sheet.cell(row=1, column=english_col_idx, value=excel_column_english)
    if chinese_col_idx is None:
        chinese_col_idx = len(headers) + 1
        sheet.cell(row=1, column=chinese_col_idx, value=excel_column_chinese)
    if tag_col_idx is None:
        tag_col_idx = len(headers) + 1
        sheet.cell(row=1, column=tag_col_idx, value=excel_column_tag)

    # 读取现有数据
    existing_english_words = set()
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=english_col_idx, max_col=english_col_idx):
        for cell in row:
            if cell.value:
                existing_english_words.add(cell.value.strip().lower())

    # 去重后的数据
    unique_data = [entry for entry in parsed_data if entry["英文"].lower() not in existing_english_words]

    # 添加新数据
    for idx, entry in enumerate(unique_data, start=sheet.max_row + 1):
        sheet.cell(row=idx, column=english_col_idx, value=entry["英文"])
        sheet.cell(row=idx, column=chinese_col_idx, value=entry["中文"])
        sheet.cell(row=idx, column=tag_col_idx, value="附肢骨及其连接")

    # 保存文件
    wb.save(excel_file_name)


# 执行代理并获取结果
result1 = agent1_extractor.execute(user_provided_text=input("请输入文本："))
print(result1)
result2 = agent2_decomposer.execute(user_provided_text=input("请输入文本："))
# 打印结果
print(result2)
process_markdown_and_insert_to_excel()
