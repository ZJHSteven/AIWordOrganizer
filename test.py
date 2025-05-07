def process_markdown_and_insert_to_excel():
    import re
    import openpyxl

    # 1. 提取 Markdown 字符串
    markdown_text = self.state.get('decomposed_markdown_list', '')

    # 如果存在 Markdown 代码块标记，则移除
    if markdown_text.startswith("```markdown\n"):
        markdown_text = markdown_text[len("```markdown\n"):]
    if markdown_text.endswith("\n```"):
        markdown_text = markdown_text[:-len("\n```")]
    markdown_text = markdown_text.strip()  # 移除首尾可能存在的空白

    # 2. 定义正则表达式
# 匹配 "数字. 单词\n" 和随后的 "   释义：... (直到下一个 "数字." 或字符串末尾)"
# (?s) 使 . 可以匹配换行符
# (\d+\.\s+.*?)\n  匹配 "数字. 单词" 这一行，捕获单词
# (\s{3}释义：.*?)(?=\n\d+\.\s+|\Z) 匹配 "   释义：" 开始的详细内容，直到下一个条目或字符串末尾
    regex = r"(\d+\.\s*(.*?)\n)(\s+释义：.*?)(?=\n\d+\.\s*|\Z)"

    parsed_data = []
    if markdown_text:
        matches = re.finditer(regex, markdown_text, re.DOTALL)
        for match in matches:
            english_word = match.group(2).strip()

            chinese_definition_and_roots = match.group(3).strip()
            chinese_definition_and_roots = chinese_definition_and_roots.replace('   ', '')  # 去除多余的空格
            chinese_definition_and_roots = chinese_definition_and_roots.replace('    ', '')

            parsed_data.append({
                "英文": english_word,
                "中文": chinese_definition_and_roots
            })

    # 3. 准备 Excel 文件路径和表头
    excel_file_name = "test_output.xlsx"  # 测试文件名
# target_excel_path = "d:\Workspace\Stable\Python\ReWordOrganizer\词根整理.xlsx" # 实际文件路径（注释掉供测试）
    excel_column_english = "单词"
    excel_column_chinese = "含义（可不填）"
    excel_column_tag = "标签（可不填，多个标签请用英文逗号分隔）"

    # 打开 Excel 文件
    wb = openpyxl.load_workbook(excel_file_name)
    sheet = wb.active

    # 获取表头
    headers = [cell.value for cell in sheet[1]]

    # 查找列索引
    english_col_idx = headers.index(excel_column_english) + 1 if excel_column_english in headers else None
    chinese_col_idx = headers.index(excel_column_chinese) + 1 if excel_column_chinese in headers else None
    tag_col_idx = headers.index(excel_column_tag) + 1 if excel_column_tag in headers else None

    # 如果表头中没有找到列，则添加表头
    if english_col_idx is None:
        english_col_idx = len(headers) + 1
        sheet.cell(row=1, column=english_col_idx, value=excel_column_english)
    if chinese_col_idx is None:
        chinese_col_idx = len(headers) + 1
        sheet.cell(row=1, column=chinese_col_idx, value=excel_column_chinese)
    if tag_col_idx is None:
        tag_col_idx = len(headers) + 1
        sheet.cell(row=1, column=tag_col_idx, value=excel_column_tag)

    # 读取现有数据
    existing_english_words = set()
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=english_col_idx, max_col=english_col_idx):
        for cell in row:
            if cell.value:
                existing_english_words.add(cell.value.strip().lower())

    # 去重后的数据
    unique_data = [entry for entry in parsed_data if entry["英文"].lower() not in existing_english_words]

    # 添加新数据
    for idx, entry in enumerate(unique_data, start=sheet.max_row + 1):
        sheet.cell(row=idx, column=english_col_idx, value=entry["英文"])
        sheet.cell(row=idx, column=chinese_col_idx, value=entry["中文"])
        sheet.cell(row=idx, column=tag_col_idx, value="系统解放式")

    # 保存文件
    wb.save(excel_file_name)



